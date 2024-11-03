from openpyxl import load_workbook, drawing
from openpyxl.styles import PatternFill, NamedStyle,Border, Side, Alignment, Font
from flask import Flask, request, jsonify, send_file
import subprocess
import os
import json
import logging
from PIL import Image


app = Flask(__name__)
os.umask(0)
@app.route('/create-excel', methods=['POST'])
def main_excel():
    content_type = request.headers.get('Content-Type')
    if (content_type == 'application/json'):
        used_data = json.loads(request.json)
        app.logger.info(used_data)
        wb = load_workbook(filename='./test.xlsx')
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        ws = add_header(ws, used_data["year"],
                            used_data["month"],
                            used_data["name"],
                            used_data["id"],
                            used_data["mail"],
                            used_data["phone"]
                            )
        ws, total_height = add_lines(ws, used_data)
        ws = sum_lines(ws, used_data, total_height)
        #img = drawing.image.Image('./logo.jpg')
        #img.anchor = 'N1'
        #ws.add_image(img)
        #img2 = drawing.image.Image('./logo.jpg')
        #img2.anchor = 'N39'
        #ws.add_image(img2)
        wb.save("/tmp/result.xlsx")
        os.chmod("/tmp/result.xlsx", 0o666)
        result = subprocess.run(["unoconv", "-f", "pdf", "/tmp/result.xlsx"], capture_output=True, text=True)
        app.logger.info(f"LibreOffice output: {result.stdout}")
        app.logger.error(f"LibreOffice error: {result.stderr}")
        try:
            return send_file('/tmp/result.pdf', as_attachment=True)
        finally:
            if os.path.exists('/tmp/result.pdf'):
                os.remove('/tmp/result.pdf')
                os.remove('/tmp/result.xlsx')
    else:
        return "Content type is not supported."


def add_header(ws, year, month, name, id, mail, phone):
    ws['C2'] = year
    ws['D2'] = month
    ws['D3'] = name
    ws['D3'].alignment = Alignment(horizontal='left')
    ws['D4'] = id
    ws['D4'].alignment = Alignment(horizontal='left')
    ws['D5'] = mail
    ws['D5'].alignment = Alignment(horizontal='left')
    ws['D6'] = phone
    ws['D6'].alignment = Alignment(horizontal='left')
    return ws


def add_lines(ws, rows):
    count = len(rows['rows'])
    total_needed = 23
    empty_border = Border()
    full_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin'),
    )
    up_down_border = Border(
        top=Side(style='thin'),  # Sadece üst kenarlık
        bottom=Side(style='thin')  # Sadece alt kenarlık
    )
    left_up_down_border = Border(
        top=Side(style='thin'),  # Sadece üst kenarlık
        bottom=Side(style='thin'),  # Sadece alt kenarlık
        left=Side(style='thin')
    )
    right_up_down_border = Border(
        top=Side(style='thin'),  # Sadece üst kenarlık
        bottom=Side(style='thin'),  # Sadece alt kenarlık
        right=Side(style='thin')
    )
    for row in range(9, total_needed + 1):
        for col in range(2, 17):
            cell = ws.cell(row=row, column=col,value=None)
            cell.border = empty_border
            cell.value = None
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    start = 9
    for index, row in enumerate(rows['rows']):
        row_values = list(row.values())
        for col in range(2, 17):
            if col == 14:
                cell = ws.cell(row=start + index, column=col)
                cell.border = full_border
                cell.value = row_values[col - 2]
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if cell.value is not None:
                    length = len(cell.value)
                    if length < 44:
                        ws.row_dimensions[start + index].height = 20
                    elif length < 88 and length > 44:
                        ws.row_dimensions[start + index].height = 40
                    elif length < 132 and length > 88:
                        ws.row_dimensions[start + index].height = 60
                    else:
                        ws.row_dimensions[start + index].height = 80

                else:
                    ws.row_dimensions[start + index].height = 20
            ws.cell(row=start + index, column=col).alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
            cell = ws.cell(row=start + index, column=col)
            cell.border = full_border
            cell.value = row_values[col - 2]

    total = str(count + start + 1)
    number_format = NamedStyle(name="number")
    number_format.number_format = '#,##0.00'
    for column in range(2,17):
        cell = ws.cell(row=int(total),column=column)
        if column == 2:
            cell.value = str(count) + " Tage"
            cell.border = left_up_down_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
        elif column == 3:
            cell.border = up_down_border
        elif column == 4:
            cell.value = rows['totals']['work_sum_amount']
            cell.style = number_format
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        elif column == 7:
            cell.value = rows['totals']['guests']
            cell.style = number_format
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        elif column == 8:
            cell.value = rows['totals']['breaks']
            cell.style = number_format
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        elif column == 9:
            cell.value = rows['totals']['public_holidays']
            cell.style = number_format
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        elif column == 10:
            cell.value = rows['totals']['sunday_holidays']
            cell.style = number_format
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        elif column == 11:
            cell.value = rows['totals']['midnight_shift']
            cell.style = number_format
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        elif column == 12:
            cell.value = rows['totals']['night_shift']
            cell.style = number_format
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        elif column == 13:
            cell.value = rows['totals']['accomodations']
            cell.fill = PatternFill(start_color="F8EEC7", end_color="F8EEC7", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = up_down_border
        else:
            cell.border = up_down_border
        if column == 16:
            cell.border = right_up_down_border
    return ws, total_needed


def write_to_excel(file_path, sheet_name, cell_coordinates, value):
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]
        sheet[cell_coordinates] = value
        wb.save(file_path)
        print(f"'{value}' değeri başarıyla '{cell_coordinates}' hücresine yazıldı.")
    except Exception as e:
        print("Excel dosyasına yazma işlemi başarısız oldu. Hata:", e)


def color_cell(file_path, sheet_name, cell_coordinates, color):
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet[cell_coordinates].fill = fill
        wb.save(file_path)
        print(f"'{cell_coordinates}' hücresi başarıyla renklendirildi.")
    except Exception as e:
        print("Excel dosyasında hücre renklendirme işlemi başarısız oldu. Hata:", e)


def sum_lines(ws,used_data, total_height):
    ws['E35'].value = str(used_data['totals']['dates']) + " Tage"
    for row in range(35, 46):
        ws.row_dimensions[row].height = 40
    ws['E36'].value = str(used_data['totals']['workhours']) + " St"
    ws['E37'].value = str(used_data['totals']['breaks']) + " St"
    ws['E38'].value = str(used_data['totals']['ausbildung_hours']) + " €"
    ws['E39'].value = str(used_data['totals']['sub_total']) + " St"
    ws['E40'].value = str(used_data['totals']['night_shift']) + " St"
    ws['E41'].value = str(used_data['totals']['midnight_shift']) + " St"
    ws['E42'].value = str(used_data['totals']['sunday_holidays']) + " St"
    ws['E43'].value = str(used_data['totals']['public_holidays']) + " St"
    ws['E44'].value = str(used_data['totals']['guests']) + " St"
    ws['E45'].value = str(used_data['totals']['total_work_day_amount']) + " €"
    ws['E46'].value = str(used_data['totals']['accomodations'])
    ws['K35'].value = str(used_data['hour_bank_this_month'])
    ws['K35'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K36'].value = str(used_data['hour_bank_this_year'])
    ws['K36'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K38'].value = str(used_data['sick_days_this_month'])
    ws['K38'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K39'].value = str(used_data['sick_days_this_year'])
    ws['K39'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K41'].value = str(used_data['annual_leave_days'])
    ws['K41'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K42'].value = str(used_data['annual_leave_rights'])
    ws['K42'].alignment = Alignment(horizontal='center', vertical='center')
    ws['K43'].value = str(used_data['annual_leave_left'])
    ws['K43'].alignment = Alignment(horizontal='center', vertical='center')
    ws['O35'].value = str(used_data['total_hours_req'])
    ws['O35'].alignment = Alignment(horizontal='center', vertical='center')
    ws['O36'].value = str(used_data['total_made_hours'])
    ws['O36'].alignment = Alignment(horizontal='center', vertical='center')
    ws['O37'].value = str(used_data['left_hours'])
    ws['O37'].alignment = Alignment(horizontal='center', vertical='center')
    ws['O37'].fill = PatternFill(start_color="f01111", end_color="f01111", fill_type="solid")
    return ws


@app.route('/create-total-excel', methods=['POST'])
def create_total_excel():
    content_type = request.headers.get('Content-Type')
    if (content_type == 'application/json'):
        used_data = json.loads(request.json)
        wb = load_workbook(filename='./total_report.xlsx')
        ws = wb.active
        rows_list = []
        for key, value in used_data['rows'].items():
            rows_list.append(value)
        used_data['rows'] = rows_list
        app.logger.info(used_data)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 25
        ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D3'] = used_data['month']
        ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D4'] = used_data['year']
        ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
        for row in used_data['rows']:
            row_num = 10
            for row_data in used_data['rows']:
                if row_num % 2 != 0:
                    for col in range(2, 19):  # B to M columns
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                for col in range(2, 19):  # B to M columns
                    cell = ws.cell(row=row_num, column=col)
                    ws.row_dimensions[row_num].height = 30
                    cell.font = Font(name='Montserrat', size=12, bold=False)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                ws[f'B{row_num}'] = f"{row_data['id']}"
                ws[f'C{row_num}'] = f"{row_data['name']}"
                ws[f'D{row_num}'] = f"{row_data['salary']}"
                ws[f'E{row_num}'] = f"{row_data['workhours']}"
                ws[f'F{row_num}'] = f"{row_data['guests']}"
                ws[f'G{row_num}'] = f"{row_data['total_work_day_amount']}"
                ws[f'H{row_num}'] = f"{row_data['ausbildung_hours']} €"
                ws[f'I{row_num}'] = f"{row_data['extra_work']}"
                ws[f'J{row_num}'] = f"{row_data['night_shift']}"
                ws[f'K{row_num}'] = f"{row_data['midnight_shift']}"
                ws[f'L{row_num}'] = f"{row_data['sunday_holidays']}"
                ws[f'M{row_num}'] = f"{row_data['public_holidays']}"
                ws[f'N{row_num}'] = f"{row_data['annual_leave_hours']}"
                ws[f'O{row_num}'] = f"{row_data['sick_leave_hours']}"
                ws[f'P{row_num}'] = f"{row_data['accomodations']}"

                if row_data['user_bonus']:
                    ws[f'Q{row_num}'] = f"+{row_data['user_bonus']} €"
                    ws[f'Q{row_num}'].font = Font(color="00FF00")
                    ws[f'Q{row_num}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                else:
                    ws[f'Q{row_num}'] = row_data['user_bonus']
                if row_data['user_advance']:
                    ws[f'R{row_num}'] = f"-{row_data['user_advance']} €"
                    ws[f'R{row_num}'].font = Font(color="FF0000")
                    ws[f'R{row_num}'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                else:
                    ws[f'R{row_num}'] = row_data['user_advance']
                row_num += 1
        wb.save("/tmp/result_total.xlsx")
        os.chmod("/tmp/result_total.xlsx", 0o666)
        result = subprocess.run(["unoconv", "-f", "pdf", "/tmp/result_total.xlsx"], capture_output=True, text=True)
        app.logger.info(f"LibreOffice output: {result.stdout}")
        app.logger.error(f"LibreOffice error: {result.stderr}")
        try:
            return send_file('/tmp/result_total.pdf', as_attachment=True)
        finally:
            if os.path.exists('/tmp/result_total.pdf'):
                os.remove('/tmp/result_total.pdf')
                os.remove('/tmp/result_total.xlsx')
    else:
        return "Content type is not supported."



@app.route('/create-excel-client-multiple', methods=['POST'])
def main_excel_client():
    content_type = request.headers.get('Content-Type')
    if (content_type == 'application/json'):
        used_data = json.loads(request.json)
        app.logger.info(used_data)
        wb = load_workbook(filename='./test_musteri.xlsx')
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        ws = add_header_client_multiple_user(ws, used_data["year"],
                            used_data["month"],
                            used_data["client"]
                            )
        ws = add_lines_client_multiple_user(ws, used_data)
        wb.save("/tmp/result_client.xlsx")

        try:
            return send_file('/tmp/result_client.xlsx', as_attachment=True)
        finally:
            if os.path.exists('/tmp/result_client.xlsx'):
                os.remove('/tmp/result_client.xlsx')
    else:
        return "Content type is not supported."

def add_header_client_multiple_user(ws, year, month, client):
    ws['C3'] = year
    ws['D3'] = month
    ws['D4'] = client
    ws['D3'].alignment = Alignment(horizontal='left', vertical='center')
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
    return ws

def add_lines_client_multiple_user(ws, rows):
    count = len(rows['rows'])
    empty_border = Border()
    full_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin'),
    )
    left_right_thick = Border(
        left=Side(style='thick'),
        top=Side(style='thin'),
        right=Side(style='thick'),
        bottom=Side(style='thin')
    )
    only_left_thick_border = Border(
        left=Side(style='thick'),
        top=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin')
    )
    up_down_border = Border(
        top=Side(style='thin'),  # Sadece st kenarlık
        bottom=Side(style='thin')  # Sadece alt kenarlık
    )
    left_up_down_border = Border(
        top=Side(style='thin'),  # Sadece üst kenarlık
        bottom=Side(style='thin'),  # Sadece alt kenarlık
        left=Side(style='thin')
    )
    right_up_down_border = Border(
        top=Side(style='thin'),  # Sadece üst kenarlık
        bottom=Side(style='thin'),  # Sadece alt kenarlık
        right=Side(style='thin')
    )
    for row in range(8, count + 8):
        ws.insert_rows(row)
        for col in range(2, 15):
            cell = ws.cell(row=row, column=col,value=None)
            cell.border = empty_border
            cell.value = None
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
    start = 8
    for index, row in enumerate(rows['rows']):
        row_values = list(row.values())
        for col in range(2, 15):
            cell = ws.cell(row=start + index, column=col)
            if(row_values[13] == 1):
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            if(row_values[14] == "true"):
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            if col == 2 or col == 4 or col == 6 or col == 8  or col == 10 or col == 11 or col == 13 or col == 14:
                cell.border = left_right_thick
            elif col == 3 or col == 5 or col == 7 or col == 9  or col == 12:
                cell.border = only_left_thick_border
            else:
                cell.border = full_border

            if col != 5 and col != 7 and col != 9:
                cell.value = row_values[col - 2]
            else:
                cell.value = row_values[col - 2] + " St"
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].auto_size = True
    total_row = count + start
    for col in range(2, 15):
        cell = ws.cell(row=total_row, column=col)
        if col == 2:
            cell.value = str(rows['totals']['total_day']) + " Tage"
            cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 5:
            cell.value = str(rows['totals']['work_total']) + " St"
            cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 7:
            cell.value = str(rows['totals']['guest_total']) + " St"
            cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 9:
            cell.value = str(rows['totals']['guest_back_total']) + " St"
            cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
    return ws

@app.route('/create-excel-client-multiple-pdf', methods=['POST'])
def main_excel_client_pdf():
    content_type = request.headers.get('Content-Type')
    if (content_type == 'application/json'):
        used_data = json.loads(request.json)
        app.logger.info(used_data)
        wb = load_workbook(filename='./test_musteri.xlsx')
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        ws = add_header_client_multiple_user(ws, used_data["year"],
                            used_data["month"],
                            used_data["client"]
                            )
        ws = add_lines_client_multiple_user(ws, used_data)

        wb.save("/tmp/result_client.xlsx")
        os.chmod("/tmp/result_client.xlsx", 0o666)
        result = subprocess.run(["unoconv", "-f", "pdf", "/tmp/result_client.xlsx"], capture_output=True, text=True)
        app.logger.info(f"LibreOffice output: {result.stdout}")
        app.logger.error(f"LibreOffice error: {result.stderr}")
        try:
            return send_file('/tmp/result_client.pdf', as_attachment=True)
        finally:
            if os.path.exists('/tmp/result_client.pdf'):
                os.remove('/tmp/result_client.pdf')
                os.remove('/tmp/result_client.xlsx')
    else:
        return "Content type is not supported."


if __name__ == "__main__":
    app.run(host="0.0.0.0",port=8000, debug=True)
